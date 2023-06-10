import os
from cv2 import COLOR_RGB2BGR
from simple_facerec import SimpleFacerec
from tkinter.messagebox import showinfo
from tkinter import *  # type: ignore
from PIL import Image, ImageTk
from tkinter import messagebox
import openpyxl as xl
from datetime import datetime
import cv2

####################################################Functions#####################################################


def saveToExcelBtn():
    wb = xl.Workbook()
    ws = wb.active
    ws.cell(1, 1, value="Date: ") # type: ignore
    ws.cell(1, 2, value=f"{datetime.now().strftime('%d-%B-%Y')}") # type: ignore
    lst = ["Name", "Enrollnment No.", "Present", "Absent"]
    for i in range(len(lst)):
        ws.cell(2, i+1, value=lst[i]) # type: ignore

    with open("./trial.csv") as fs:
        fs.seek(0)
        lst_names = [str(i).split(",")[0] for i in fs.readlines()]
        fs.seek(0)
        lst_enroll = [str(i[:len(i)-1]).split(",")[1] for i in fs.readlines()]
    for i in range(len(lst_names)):
        ws.cell(i+3, 1, value=lst_names[i]) # type: ignore
        if i in lb3.curselection():
            ws.cell(i+3, 3, value="P") # type: ignore
        else:
            ws.cell(i+3, 4, value="A") # type: ignore
    for i in range(len(lst_enroll)):
        ws.cell(i+3, 2, value=lst_enroll[i]) # type: ignore
    # for i in lb3.curselection():
    #     print(lb3.index(i))
    #     print(lb3.get(type(i)))
    wb.save(
        f"./Attendence Sheets/BCA {rb.get()} Attendence {datetime.now().strftime('%d-%B-%Y')}.xlsx")


def openFileBtn():
    os.startfile(
        f'''.\\Attendence Sheets\\{lb.get(lb.curselection())}''')


def loadImageBtn():
    lb2.delete(0, END)
    for i in os.listdir("./photos/"):
        lb2.insert(END, str(i))


def captureBtn():
    l2.grid_forget()
    cv2.imwrite("./temp.jpg", cap.read()[1])
    img2 = ImageTk.PhotoImage(Image.open("./temp.jpg").resize(  # type: ignore
        (root.winfo_screenheight() // (root.winfo_screenheight() // 400),
         root.winfo_screenwidth() // (root.winfo_screenwidth() // 400))))
    l3.configure(image=img2)
    l3.image = img2  # type: ignore
    l3.grid(row=1, rowspan=3, column=2, pady=50)
    cap.release()


def selectImageBtn():
    img1 = ImageTk.PhotoImage(Image.open(f"./photos/{lb2.get(lb2.curselection())}").resize(  # type: ignore  # type: ignore
        (root.winfo_screenheight() // 4, root.winfo_screenwidth() // 4)))  # type: ignore
    l1.config(image=img1)
    l1.image = img1  # type: ignore


def sendMailBtn():
    pass


def useWebCamBtn():
    f1.pack_forget()
    f2.pack_forget()
    f3.pack_forget()
    f4.pack_forget()
    f5.pack(fill=BOTH, expand=1)


def train_faceBtn():
    f1.pack_forget()
    f3.pack_forget()
    f4.pack_forget()
    f5.pack_forget()
    f2.pack(fill=BOTH, expand=1)


def manuallyAttendenceBtn():
    f1.pack_forget()
    f2.pack_forget()
    f4.pack_forget()
    f5.pack_forget()
    f3.pack(fill=BOTH, expand=1)


def openCameraBtn():
    f1.pack_forget()
    f2.pack_forget()
    f3.pack_forget()
    f5.pack_forget()
    f4.pack(fill=BOTH, expand=1)


def attendenceSheetBtn():
    lb.delete(0, END)
    for i in os.listdir("./Attendence Sheets/"):
        lb.insert(END, str(i))


def backBtn():
    f1.pack(fill=BOTH, expand=1)
    f2.pack_forget()
    f3.pack_forget()
    f4.pack_forget()
    f5.pack_forget()

def submit_Btn1():
    print("submitbtn1")
    if name1.get() == "" or erl_no1.get() == "":
        messagebox.showerror("Name or Enrollment number not found!", "Please Enter Name and Erollnment Number")
    else:
        img3 = Image.open(f"./photos/{lb2.get(lb2.curselection())}")  # type: ignore # type: ignore
        img3.save(f"./images/{name1.get()} {erl_no1.get()}.jpg")
        os.remove(f"./photos/{lb2.get(lb2.curselection())}")

def submit_Btn2():
    print("submitbtn2")
    if name2.get() == "" or erl_no2.get() == "":
        messagebox.showerror("Name or Enrollment number not found!", "Please Enter Name and Erollnment Number")
    else:
        img3 = Image.open("./temp.jpg")  # type: ignore
        img3.save(f"./photos/{name2.get()} {erl_no2.get()}.jpg")
        os.remove("./temp.jpg")
        messagebox.showinfo("Successuful", "Your Details has been")


def clear_Btn():
    name1.set("")
    erl_no1.set("")
    name2.set("")
    erl_no2.set("")
    


def radioBtn1():
    lb3.delete(0, END)
    sc3 = Scrollbar(f3, orient=VERTICAL)
    sc3.grid(
        row=1,
        column=2,
        columnspan=1,
        sticky=E,
        ipady=root.winfo_screenheight() // (root.winfo_screenheight() // 225),
        rowspan=3,
    )
    with open("./trial.csv") as fs:
        for i in fs.readlines():
            lb3.insert(END, str(i))
    lb3.config(yscrollcommand=sc3.set)
    sc3.config(command=lb3.yview)
    # print(rb.get())


def openCamBtn1():
    cv2image = cv2.cvtColor(cap.read()[1], cv2.COLOR_BGR2RGB)
    img = Image.fromarray(cv2image)  # type: ignore
    imgtk = ImageTk.PhotoImage(image=img.resize((800,500)))
    l2.imgtk = imgtk  # type: ignore
    l2.configure(image=imgtk)
    l2.after(20, openCamBtn1)


def openCamBtn2():
    ret, frame = cap2.read()
    sfr.load_encoding_images("photos\\")
    try:
        # Detect Faces
        face_locations, face_names = sfr.detect_known_faces(frame)
        for face_loc, name in zip(face_locations, face_names):
            y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]

            if name!="Unknown":
                cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (54, 240, 7), 1)
                cv2.rectangle(frame, (x1, y1), (x2, y2), (54, 240, 7), 1)
            else:
                cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 200), 1)
                cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 200), 1)

        bgrframe = cv2.cvtColor(frame,COLOR_RGB2BGR)
        img = Image.fromarray(bgrframe).resize((800,500))  # type: ignore
        imgtk = ImageTk.PhotoImage(image = img)
        l4.config(image=imgtk)
        l4.image = imgtk  # type: ignore
        with open("./Attendence Sheets/Attendence.csv", 'r+') as f:
            myDataList = f.readlines()
            nameList = []
            for line in myDataList:
                entry = line.split(',')
                nameList.append(entry[0])
            for name in face_names:
                if name not in nameList:
                    now = datetime.now()
                    time = now.strftime('%I:%M:%S:%p')
                    date = now.strftime('%d-%B-%Y')
                    f.writelines(f'{name}, {time}, {date}\n')
    except:
        pass
    l4.after(10,openCamBtn2)


def closeCamBtn():
    cap2.release()
    l4.config(image=img3)

def openAttSheetBtn():
    os.startfile(".\\Attendence Sheets\\Attendence.csv")

#####################################################Main######################################################
root = Tk()
root.config(background="White")
root.state("zoomed")
print(root.winfo_screenheight())
print(root.winfo_screenwidth())
f1 = Frame(root, bg="White")
f2 = Frame(root, bg="White")
f3 = Frame(root, bg="White")
f4 = Frame(root, bg="White")
f5 = Frame(root, bg="White")
sfr = SimpleFacerec()
sfr.load_encoding_images("photos\\")

#####################################################Frame 1###################################################
btn_lst = ["Attendence Sheets", "Train Face",
           "Manually Attendence", "Open Webcam"]
place_lst = [
    i for i in range(0, root.winfo_screenwidth(), root.winfo_screenwidth() // 4)
]
func_lst = [attendenceSheetBtn, train_faceBtn,
            manuallyAttendenceBtn, openCameraBtn]
for i in range(len(btn_lst)):
    Button(
        f1,
        background="black",
        fg="White",
        height=root.winfo_screenheight() // 146,
        width=root.winfo_screenwidth() // 41,
        text=btn_lst[i],
        font=("Helvetica", 13, "bold"),
        command=func_lst[i],
    ).grid(row=0, column=i)
lb = Listbox(
    f1,
    bg="white",
    height=root.winfo_screenheight() // 51,
    width=root.winfo_screenwidth() // 21,
    bd=5,
    relief=RIDGE,
    font=("Helvetica", 16),
    setgrid=True,
    selectbackground="black",
)
lb.grid(
    row=2,
    columnspan=4,
    pady=root.winfo_screenheight() // (root.winfo_screenheight() // 100),
)
Button(
    f1,
    text="Open File",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // 341,
    width=root.winfo_screenwidth() // 106,
    relief=RIDGE,
    command=openFileBtn,
).grid(row=3, column=1, sticky=NW)

f1.pack(fill=BOTH, expand=1)

#####################################################Frame 2###################################################
Label(
    f2,
    bg="black",
    text="Train Face",
    fg="White",
    font=("Helvetica", 20),
    height=root.winfo_screenheight() // 200,
    width=root.winfo_screenwidth() // 16,
).grid(row=0, columnspan=3, sticky=EW)
img1 = ImageTk.PhotoImage(
    Image.open("./nothing.jpg").resize(
        (root.winfo_screenheight() // 4, root.winfo_screenwidth() // 4)
    )
)
l1 = Label(f2, bd=7, image=img1, background="black", height=root.winfo_screenheight() //
           (root.winfo_screenheight() // 300), width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 300), relief=RIDGE)
l1.grid(row=1, rowspan=3, column=0, pady=50)
Label(f2, text="Name : ", font=("Helvetica", 16), bg="White").grid(
    row=1, column=1, sticky=SW, columnspan=2
)
Label(f2, text="Enrollment No. : ", font=("Helvetica", 16), bg="White").grid(
    row=2, column=1, sticky=W, columnspan=3
)
name1 = StringVar(f2)
erl_no1 = StringVar(f2)
Entry(
    f2, relief=RIDGE, textvariable=name1, width=root.winfo_screenwidth() // 25, bd=3
).grid(row=1, column=1, sticky=SE)
Entry(
    f2, relief=RIDGE, textvariable=erl_no1, width=root.winfo_screenwidth() // 25, bd=3
).grid(row=2, column=1, sticky=E)
Button(
    f2,
    text="Back",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // 341,
    width=root.winfo_screenwidth() // 106,
    relief=RIDGE,
    command=backBtn,
).grid(row=3, column=1, sticky=NW)
Button(
    f2,
    text="Submit",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // 341,
    width=root.winfo_screenwidth() // 106,
    relief=RIDGE,
    command=submit_Btn1,
).grid(row=3, column=1, sticky=N)
Button(
    f2,
    text="Clear",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // 341,
    width=root.winfo_screenwidth() // 106,
    relief=RIDGE,
    command=clear_Btn,
).grid(row=3, column=1, sticky=NE)
lb2 = Listbox(
    f2,
    bg="white",
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 13),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 30),
    bd=5,
    relief=RIDGE,
    font=("Helvetica", 16),
    setgrid=True,
    selectbackground="black",
)
lb2.grid(
    row=4,
    rowspan=3,
    column=1,
    columnspan=3,
    sticky=W,
    padx=root.winfo_screenwidth() // 25,
)
Button(
    f2,
    text="Load Image",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // 341,
    width=root.winfo_screenwidth() // 106,
    relief=RIDGE,
    command=loadImageBtn,
).grid(row=4, column=0)
Button(
    f2,
    text="Select Image",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // 341,
    width=root.winfo_screenwidth() // 106,
    relief=RIDGE,
    command=selectImageBtn,
).grid(row=5, column=0)

Button(
    f2,
    text="Use Webcam",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // 341,
    width=root.winfo_screenwidth() // 106,
    relief=RIDGE,
    command=useWebCamBtn
).grid(row=6, column=0)

# f2.pack(fill=BOTH, expand=1)

#####################################################Frame 3###################################################
Label(
    f3,
    bg="black",
    text="Manual Attendence",
    fg="White",
    font=("Helvetica", 20),
    height=root.winfo_screenheight() // 200,
    width=root.winfo_screenwidth() // 16,
).grid(row=0, columnspan=3, sticky=EW)
rb = IntVar()

count = 1
for i in range(1, 4):
    for j in range(2):
        if j == 0:
            Radiobutton(
                f3,
                bd=3,
                background="white",
                text=f"BCA {count} Sem",
                font=("Helvetica", 13),
                activebackground="white",
                selectcolor="white",
                variable=rb,
                value=count,
                command=radioBtn1,
            ).grid(
                row=i,
                column=j,
                sticky=N,
                pady=root.winfo_screenheight() // (root.winfo_screenheight() // 60),
            )
        else:
            Radiobutton(
                f3,
                bd=3,
                background="white",
                text=f"BCA {count} Sem",
                font=("Helvetica", 13),
                activebackground="white",
                selectcolor="white",
                variable=rb,
                value=count,
                command=radioBtn1,
            ).grid(
                row=i,
                column=j,
                sticky=NW,
                pady=root.winfo_screenheight() // (root.winfo_screenheight() // 60),
            )
        count += 1

Button(
    f3,
    text="Back",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=backBtn,
).grid(row=4, column=0)
Button(
    f3,
    text="Send Mail",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=sendMailBtn,
).grid(row=4, column=1, sticky=W)
Button(
    f3,
    text="Save to Excel",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=saveToExcelBtn,
).grid(row=4, column=2, sticky=W)
lb3 = Listbox(
    f3,
    bg="white",
    height=20,
    width=30,
    bd=5,
    relief=RIDGE,
    font=("Helvetica", 16),
    setgrid=True,
    selectbackground="black",
    selectmode=MULTIPLE,
)
lb3.grid(row=1, column=2, sticky=W, pady=60, rowspan=3)
# f3.pack(fill=BOTH, expand=1)

#####################################################Frame 4###################################################
Label(
    f4,
    bg="black",
    text="Open Webcam",
    fg="White",
    font=("Helvetica", 20),
    height=root.winfo_screenheight() // 200,
    width=root.winfo_screenwidth() // 16,
).grid(row=0, columnspan=3, sticky=EW)
Button(
    f4,
    text="Open Camera",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 15),
    relief=RIDGE,
    command=openCamBtn2,
).grid(row=1, column=1)
img3 = ImageTk.PhotoImage(
    Image.open("./nothing.jpg").resize(
        (root.winfo_screenheight() // 4, root.winfo_screenwidth() // 4)
    )
)
l4 = Label(f4, bd=7, image=img3, background="black", height=root.winfo_screenheight() //
           (root.winfo_screenheight() // 500), width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 700), relief=RIDGE)
l4.grid(row=1, rowspan=3, column=2, pady=50)
cap2 = cv2.VideoCapture(0)
Button(
    f4,
    text="Close Camera",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 15),
    relief=RIDGE,
    command=closeCamBtn,
).grid(row=1, column=1, sticky=S)
Button(
    f4,
    text="Open Attendence Sheet",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 25),
    relief=RIDGE,
    command=openAttSheetBtn,
).grid(row=2, column=1)

Button(
    f4,
    text="Back",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=backBtn,
).grid(row=4, column=0)

# f4.pack(fill=BOTH, expand=1)
#####################################################Frame 5###################################################
Label(
    f5,
    bg="black",
    text="Open Webcam",
    fg="White",
    font=("Helvetica", 20),
    height=root.winfo_screenheight() // 200,
    width=root.winfo_screenwidth() // 16,
).grid(row=0, columnspan=3, sticky=EW)
Button(
    f5,
    text="Open Camera",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=openCamBtn1,
).grid(row=1, column=1, sticky=S)
img2 = ImageTk.PhotoImage(
    Image.open("./nothing.jpg").resize(
        (root.winfo_screenheight() // 4, root.winfo_screenwidth() // 4)
    )
)
l2 = Label(f5, bd=7, image=img2, background="black", height=root.winfo_screenheight() //
           (root.winfo_screenheight() // 400), width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 430), relief=RIDGE)
l3 = Label(f5, bd=7, image=img2, background="black", height=root.winfo_screenheight() //
           (root.winfo_screenheight() // 400), width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 430), relief=RIDGE)
cap = cv2.VideoCapture(0)
l2.grid(row=1, rowspan=3, column=2, pady=50)

Label(f5, text="Name : ", font=("Helvetica", 16), bg="White").grid(
    row=2, column=0, sticky=S
)
Label(f5, text="Enrollment No. : ", font=("Helvetica", 16), bg="White").grid(
    row=3, column=0
)
name2 = StringVar(f5)
erl_no2 = StringVar(f5)
Entry(
    f5, relief=RIDGE, textvariable=name2, font=("Helvetica", 16), width=root.winfo_screenwidth() // 25, bd=3
).grid(row=2, column=1, sticky=SE)
Entry(
    f5, relief=RIDGE, textvariable=erl_no2, font=("Helvetica", 16), width=root.winfo_screenwidth() // 25, bd=3
).grid(row=3, column=1, sticky=E)
Button(
    f5,
    text="Capture Photo",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=captureBtn,
).grid(row=4, column=2)
Button(
    f5,
    text="Back",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=backBtn,
).grid(row=4, column=1, sticky=SW)
Button(
    f5,
    text="Submit",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=submit_Btn2,
).grid(row=4, column=1, sticky=S)
Button(
    f5,
    text="Clear",
    bg="Black",
    fg="white",
    font=("Helvetica", 13),
    height=root.winfo_screenheight() // (root.winfo_screenheight() // 3),
    width=root.winfo_screenwidth() // (root.winfo_screenwidth() // 12),
    relief=RIDGE,
    command=clear_Btn,
).grid(row=4, column=1, sticky=E)
# f5.pack(fill=BOTH, expand=1)


root.mainloop()
